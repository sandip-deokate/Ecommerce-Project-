from multiprocessing.pool import worker

import openpyxl
import os
class ExcelOperations:
    def __init__(self,driver):
        self.driver=driver
    """path= os.path.abspath(os.curdir+ '//testData//'+'BomShort.xlsx')
    workbook=openpyxl.load_workbook("path")
    sheet= workbook["Best Buying Options"]

    max_row= sheet.max_row
    max_col=sheet.max_column"""

    def get_max_row(self,path,sheetname):
        workbook=openpyxl.load_workbook(path)
        sheet=workbook[sheetname]
        return sheet.max_row

    def get_max_col(self,path,sheetname):
        workbook=openpyxl.load_workbook(path)
        sheet=workbook[sheetname]
        return sheet.max_column

    def read_data(self,path,sheetname,row,col):
        workbook=openpyxl.load_workbook(path)
        sheet=workbook[sheetname]
        return sheet.cell(row,col).value

    def write_Data(self,path,sheetname,row,col,data):
        workbook=openpyxl.load_workbook(row,col)
        sheet=workbook[sheetname]
        sheet.cell(row,col).value=data
        workbook.save(path)
        return sheet.cell(row,col).value




